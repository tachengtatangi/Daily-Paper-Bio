#!/usr/bin/env python3
"""Load and query CAS quartile workbook data."""

from __future__ import annotations

import re
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook


DATA_PATH = Path(__file__).resolve().parent / "data" / "cas_quartiles_2025.xlsx"
SHEET_NAME = "完整版"


def normalize_journal_name(name: str) -> str:
    text = (name or "").strip().upper()
    text = re.sub(r"\([^)]*\)", " ", text)
    text = text.replace("&", "AND")
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


@lru_cache(maxsize=1)
def load_quartile_map() -> dict[str, dict]:
    if not DATA_PATH.exists():
        return {}

    wb = load_workbook(DATA_PATH, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        return {}

    ws = wb[SHEET_NAME]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return {}

    header_index = {str(col).strip(): idx for idx, col in enumerate(header) if col is not None}
    journal_idx = header_index.get("期刊名称")
    quartile_idx = header_index.get("2025分区")
    top_idx = header_index.get("Top")
    oa_idx = header_index.get("Open Access")
    if journal_idx is None or quartile_idx is None:
        return {}

    result: dict[str, dict] = {}
    for row in rows:
        if not row:
            continue
        journal = row[journal_idx] if journal_idx < len(row) else None
        quartile = row[quartile_idx] if quartile_idx < len(row) else None
        if not journal or quartile in (None, ""):
            continue
        key = normalize_journal_name(str(journal))
        if not key:
            continue
        try:
            q_value = int(quartile)
        except Exception:
            continue
        result[key] = {
            "journal": str(journal).strip(),
            "quartile": q_value,
            "top": str(row[top_idx]).strip() if top_idx is not None and top_idx < len(row) and row[top_idx] is not None else "",
            "open_access": str(row[oa_idx]).strip() if oa_idx is not None and oa_idx < len(row) and row[oa_idx] is not None else "",
        }
    return result


def lookup_journal(journal_name: str) -> dict | None:
    key = normalize_journal_name(journal_name)
    if not key:
        return None
    return load_quartile_map().get(key)
